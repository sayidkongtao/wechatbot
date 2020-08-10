# coding=utf-8
import sys
import numpy as np
import cv2
from appium import webdriver
from appium.webdriver.common.multi_action import MultiAction
import time
import os
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from appium.webdriver.common.mobileby import MobileBy
from appium.webdriver.common.touch_action import TouchAction
from openpyxl import load_workbook
import logging
import re
from urllib import unquote
from openpyxl.styles import PatternFill

reload(sys)
sys.setdefaultencoding('utf8')

PATH = lambda path: os.path.abspath(
    os.path.join(
        os.path.dirname(__file__),
        path
    )
)

# 创建一个logging对象
logger = logging.getLogger()
logger.setLevel(level=logging.INFO)
# 创建一个文件对象  创建一个文件对象,以UTF-8 的形式写入 标配版.log 文件中
fh = logging.FileHandler(PATH(os.path.join("log", "example.log")), mode="w", encoding='utf-8')
# 创建一个屏幕对象
sh = logging.StreamHandler()
# 配置显示格式  可以设置两个配置格式  分别绑定到文件和屏幕上
formatter = logging.Formatter('%(asctime)s %(filename)s[line:%(lineno)d] %(levelname)s %(message)s')
fh.setFormatter(formatter)  # 将格式绑定到两个对象上
sh.setFormatter(formatter)

logger.addHandler(fh)  # 将两个句柄绑定到logger
logger.addHandler(sh)


class Utils:

    def __init__(self):
        pass

    @staticmethod
    def get_tap_coordinate(source_path, template_path):
        # todo: need to update function
        Utils.match_image_by_flann_func(source_path, template_path)

    @staticmethod
    def match_image_by_match_template_func(source_path, template_path, mobile_window_rect=None,
                                           method=cv2.TM_CCOEFF_NORMED):
        """

        :param source_path:
        :param template_path:
        :param method:
        :return:
        """
        source = cv2.imread(source_path)
        template = cv2.imread(template_path, 0)
        source_gray = cv2.cvtColor(source, cv2.COLOR_BGR2GRAY)
        result = cv2.matchTemplate(source_gray, template, method)
        w, h = template.shape[::-1]

        threshold = 0.8
        loc = np.where(result >= threshold)

        final_location = []

        for pt in zip(*loc[::-1]):
            # cv2.rectangle(img_rgb, pt, (pt[0] + w, pt[1] + h), (0, 0, 255), 1)
            final_location.append({"x": pt[0], "y": pt[1], "width": w, "height": h})

        if len(final_location) > 0:
            x = final_location[-1]["x"] + final_location[-1]["width"] / 3
            y = final_location[-1]["y"] + final_location[-1]["height"] / 3

            if mobile_window_rect:
                mobile_window_width = mobile_window_rect["width"]
                mobile_window_height = mobile_window_rect["height"]
                h, w, c = source.shape
                x = x * mobile_window_width / w
                y = y * mobile_window_height / h
            logger.info("Get the location by compare the screenshot: {}".format({"x": x, "y": y}))
            return {"x": x, "y": y}
        else:
            return None

    @staticmethod
    def match_image_by_flann_func(source_path, template_path, mobile_window_rect=None):
        """

        :param source_path:
        :param template_path:
        :param mobile_view_rect: {u'y': 0, u'width': 375, u'x': 0, u'height': 667}
        :return:
        """
        query_image = cv2.imread(PATH(template_path), 0)
        # 读取要匹配的灰度照片
        training_image = cv2.imread(PATH(source_path), 0)

        MIN_MATCH_COUNT = 10  # 设置最低特征点匹配数量为10
        # Initiate SIFT detector创建sift检测器
        sift = cv2.xfeatures2d.SIFT_create()
        # find the keypoints and descriptors with SIFT
        kp1, des1 = sift.detectAndCompute(query_image, None)
        kp2, des2 = sift.detectAndCompute(training_image, None)
        # 创建设置FLANN匹配
        FLANN_INDEX_KDTREE = 0
        index_params = dict(algorithm=FLANN_INDEX_KDTREE, trees=5)
        search_params = dict(checks=50)
        flann = cv2.FlannBasedMatcher(index_params, search_params)
        matches = flann.knnMatch(des1, des2, k=2)
        # store all the good matches as per Lowe's ratio test.
        good = []
        # 舍弃小于0.7的匹配
        for m, n in matches:
            if m.distance > 0.7 * n.distance:
                good.append(m)

        if len(good) > MIN_MATCH_COUNT:
            # 获取关键点的坐标
            src_pts = np.float32([kp1[m.queryIdx].pt for m in good]).reshape(-1, 1, 2)
            dst_pts = np.float32([kp2[m.trainIdx].pt for m in good]).reshape(-1, 1, 2)
            # 计算变换矩阵和MASK
            M, mask = cv2.findHomography(src_pts, dst_pts, cv2.RANSAC, 5.0)
            # matchesMask = mask.ravel().tolist()
            h, w = query_image.shape
            # 使用得到的变换矩阵对原图像的四个角进行变换，获得在目标图像上对应的坐标
            pts = np.float32([[0, 0], [0, h - 1], [w - 1, h - 1], [w - 1, 0]]).reshape(-1, 1, 2)
            dst = cv2.perspectiveTransform(pts, M)
            loc = (np.int32(dst[0][0]), np.int32(dst[2][0]))
            x = (np.int32(dst[2][0])[0] + np.int32(dst[0][0])[0]) / 3
            y = (np.int32(dst[2][0])[1] + np.int32(dst[0][0])[1]) / 3

            if mobile_window_rect:
                mobile_window_width = mobile_window_rect["width"]
                mobile_window_height = mobile_window_rect["height"]
                h, w = training_image.shape
                x = x * mobile_window_width / w
                y = y * mobile_window_height / h

            return {"x": x, "y": y}
        else:
            logger.error("Not enough matches are found - %d/%d" % (len(good), MIN_MATCH_COUNT))
            return None

    @staticmethod
    def load_data_from_excel(file_name):
        """

        :param file_name:
        :return:
        """
        wb = load_workbook(file_name)
        work_sheet = wb["data"]
        test_data = []
        flag_count = 0
        for row in work_sheet.values:
            if flag_count == 0:
                test_data.append(CaseDataModel(row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8],
                                               row[9], row[10]))
                flag_count = flag_count + 1
            else:
                test_data.append(CaseDataModel(row[0], row[1], row[2], row[3], row[4], None, None, None, None, None,
                                               None))

        config_sheet = wb["config"]
        config_data = {"os": config_sheet["A2"].value, "name": config_sheet["B2"].value}
        wb.close()

        return {"test_data": test_data, "config_data": config_data}

    @staticmethod
    def write_data_into_excel(file_name, data_list, start_row=2):
        """

        :param file_name:
        :param data_list:
        :param start_row:
        :return:
        """
        red_fill = PatternFill("solid", fgColor="FF0000")
        green_fill = PatternFill("solid", fgColor="00008000")
        wb = load_workbook(file_name)
        work_sheet = wb["data"]
        for data in data_list:
            work_sheet["F" + str(start_row)].value = data.reply_from_script
            work_sheet["G" + str(start_row)].value = data.link_from_script
            work_sheet["H" + str(start_row)].value = data.screenshot_from_script
            work_sheet["I" + str(start_row)].value = data.link_screenshot_from_script
            work_sheet["J" + str(start_row)].value = data.reply_cost_time_from_script
            if data.reply_from_script == data.reply:
                work_sheet["K" + str(start_row)].value = "Pass"
                work_sheet["K" + str(start_row)].fill = green_fill
            else:
                work_sheet["K" + str(start_row)].value = "Failed"
                work_sheet["K" + str(start_row)].fill = red_fill

            start_row = start_row + 1
        wb.save(file_name)


class MobileFunction:

    def __init__(self, driver):
        """

        :param driver:
        """
        self.driver = driver
        self.webdriver_wait = WebDriverWait(self.driver, 30)

    @staticmethod
    def click(element):
        element.click()

    @staticmethod
    def send_text(element, value):
        element.send_keys(value)

    @staticmethod
    def get_text(element):
        return element.text

    def wait_for_element_visible(self, locator):
        return self.webdriver_wait.until(EC.visibility_of_element_located(locator))

    def wait_for_element_presence(self, locator):
        return self.webdriver_wait.until(EC.presence_of_element_located(locator))

    def is_element_visible(self, locator):
        try:
            return self.driver.find_element(locator[0], locator[1])
        except Exception as err:
            return False

    def find_elements(self, locator):
        return self.driver.find_elements(locator[0], locator[1])

    @staticmethod
    def get_rect(element):
        return element.rect

    def right_or_left(self, element):
        pass

    def tap(self, x, y):
        touch_action = TouchAction(self.driver)
        touch_action.tap(x=int(x), y=int(y)).release().perform()

    def double_tap(self, x, y):
        action1 = TouchAction(self.driver).long_press(x=x, y=y, duration=500).release()
        action2 = TouchAction(self.driver).long_press(x=x, y=y, duration=500).release()
        m_action = MultiAction(self.driver)
        m_action.add(action1, action2)
        m_action.perform()

    def double_tap_ele_to_get_details_message(self, x, y, try_count=60):
        message = None
        touch_action = TouchAction(self.driver)
        logger.info("The details_message coordinate is {} {}".format(x, y))
        current = time.time()
        while time.time() - current < try_count:
            touch_action.tap(x=x, y=y, count=2).release().perform()
            time.sleep(0.5)
            ele = self.is_element_visible(AndroidMobilePageObject.details_message())
            if ele:
                logger.info("Try to get details message")
                message = self.get_text(ele)
                logger.info(message)
                self.click(ele)
                break

        if message:
            logger.info("Success to get the details message")
        else:
            logger.error(
                "Failed to get the details message after 60s, since cannot get the messgae by double click element")

        return message

    def get_details_message_by_copy_past_item(self, x, y, copy_item_path, time_out=60):
        message = None
        touch_action = TouchAction(self.driver)
        current = time.time()
        logger.info("信息的坐标是 x = {}, y = {}".format(x, y))
        while time.time() - current < time_out:
            touch_action.long_press(x=int(x), y=int(y), duration=500).release().perform()
            time.sleep(1)

            current_screenshot = os.path.join(copy_item_path, "current_screenshot_copy.png")
            self.driver.save_screenshot(current_screenshot)
            time.sleep(1)

            loc = Utils.match_image_by_match_template_func(current_screenshot, os.path.join(
                copy_item_path,
                "copy_item.png"
            ))

            if loc:
                self.tap(loc["x"], loc["y"])
                time.sleep(1)
                message = self.driver.get_clipboard_text()
                break

        if message:
            logger.info("成功获取到信息: {}".format(message))
        else:
            logger.error("在60s内获取信息失败")

        return message

    def save_screenshot_as_png(self, file_name):
        self.driver.save_screenshot(filename=file_name)
        logger.info("Capture the screen: {}".format(file_name))


class AndroidMobilePageObject:

    def __init__(self):
        pass

    @staticmethod
    def search_btn_in_home_page():
        return (MobileBy.ID, "com.tencent.mm:id/f8y")

    @staticmethod
    def address_book_in_home_page():
        return (MobileBy.XPATH, "(//*[@resource-id='com.tencent.mm:id/cn_'])[2]")

    @staticmethod
    def gongzong_number_item():
        return (MobileBy.XPATH, "(//*[@resource-id='com.tencent.mm:id/b3b'])[4]")

    @staticmethod
    def search_in_gongzong_page():
        return (MobileBy.ACCESSIBILITY_ID, "搜索")

    @staticmethod
    def search_input_in_gongzong_page():
        return (MobileBy.ID, "com.tencent.mm:id/bhn")

    @staticmethod
    def target_item():
        return (MobileBy.XPATH, "//android.widget.TextView[@text='大众汽车金融中国测试号']")

    @staticmethod
    def title_in_chat():
        return (MobileBy.ID, "com.tencent.mm:id/gas")

    @staticmethod
    def message_btn():
        return (MobileBy.ID, "com.tencent.mm:id/aly")

    @staticmethod
    def message_input():
        return (MobileBy.ID, "com.tencent.mm:id/al_")

    @staticmethod
    def message_send_btn():
        return (MobileBy.ID, "com.tencent.mm:id/anv")

    # 服务按钮 com.tencent.mm:id/g33
    @staticmethod
    def menu_btn():
        return (MobileBy.ID, "com.tencent.mm:id/g33")

    @staticmethod
    def latest_message():
        return (MobileBy.ID, "com.tencent.mm:id/ala")

    @staticmethod
    def details_message():
        return (MobileBy.ID, "com.tencent.mm:id/c9a")

    @staticmethod
    def contact_photo():
        return (MobileBy.ID, "com.tencent.mm:id/aku")

    @staticmethod
    def back_btn():
        return (MobileBy.ID, "com.tencent.mm:id/dn")


class IOSMobilePageObject:

    def __init__(self):
        pass

    @staticmethod
    def address_book_in_home_page():
        return (MobileBy.ACCESSIBILITY_ID, "通讯录")

    @staticmethod
    def gongzong_number_item():
        return (MobileBy.ACCESSIBILITY_ID, "公众号")

    @staticmethod
    def search_in_gongzong_page():
        return (MobileBy.ACCESSIBILITY_ID, "搜索")

    @staticmethod
    def search_input_in_gongzong_page():
        return (MobileBy.ACCESSIBILITY_ID, "搜索")

    @staticmethod
    def target_item():
        return (MobileBy.XPATH, "(//XCUIElementTypeStaticText[@name='大众汽车金融中国测试号'])[3]")

    @staticmethod
    def title_in_chat():
        return (MobileBy.ACCESSIBILITY_ID, "聊天详情")

    @staticmethod
    def message_btn():
        return (MobileBy.ACCESSIBILITY_ID, "切换到文本输入")

    @staticmethod
    def message_input():
        return (MobileBy.XPATH, "//XCUIElementTypeTextView")

    @staticmethod
    def message_send_btn():
        return (MobileBy.ACCESSIBILITY_ID, "Send")

    # 切换到菜单
    @staticmethod
    def menu_btn():
        return (MobileBy.ACCESSIBILITY_ID, "切换到菜单")

    @staticmethod
    def latest_message():
        return (MobileBy.XPATH, "//XCUIElementTypeTable//XCUIElementTypeCell[last()]//XCUIElementTypeOther")

    @staticmethod
    def all_message():
        return (MobileBy.XPATH, "//XCUIElementTypeTable//XCUIElementTypeCell//XCUIElementTypeOther")

    @staticmethod
    def back_btn():
        return (MobileBy.ACCESSIBILITY_ID, "关闭")


class AndroidProcess:

    def __init__(self, webdriver):
        self.driver = webdriver
        self.send_message_time = None
        self.get_reply_time = None
        self.mobile_function = MobileFunction(self.driver)

    def go_into_volkswagen_official_account(self, count_name):
        ele_address_book_btn = self.mobile_function.wait_for_element_visible(
            AndroidMobilePageObject.address_book_in_home_page()
        )
        self.mobile_function.click(ele_address_book_btn)

        ele_gongzong_number_item = self.mobile_function.wait_for_element_presence(
            AndroidMobilePageObject.gongzong_number_item()
        )
        self.mobile_function.click(ele_gongzong_number_item)

        ele_search_in_gongzong_page = self.mobile_function.wait_for_element_visible(
            AndroidMobilePageObject.search_in_gongzong_page()
        )
        self.mobile_function.click(ele_search_in_gongzong_page)

        ele_search_input_in_gongzong_page = self.mobile_function.wait_for_element_visible(
            AndroidMobilePageObject.search_input_in_gongzong_page()
        )
        # "大众汽车金融中国测试号"
        self.mobile_function.send_text(ele_search_input_in_gongzong_page, count_name.decode("utf-8"))

        ele_target_item = self.mobile_function.wait_for_element_visible(
            AndroidMobilePageObject.target_item()
        )
        self.mobile_function.click(ele_target_item)

        # ele_title_in_chat = self.mobile_function.wait_for_element_visible(AndroidMobilePageObject.title_in_chat())

    def send_message_then_calculating_time_taken_to_reply(self, value):
        value = value.decode("utf-8")
        logger.info("发送信息: " + value)
        # self.mobile_function.wait_for_element_visible(AndroidMobilePageObject.title_in_chat())
        time.sleep(1)
        ele_message = self.mobile_function.is_element_visible(AndroidMobilePageObject.message_btn())
        if ele_message:
            self.mobile_function.click(ele_message)

        ele_message_input = self.mobile_function.wait_for_element_visible(AndroidMobilePageObject.message_input())
        self.mobile_function.send_text(ele_message_input, value)
        ele_message_send_btn = self.mobile_function.wait_for_element_visible(AndroidMobilePageObject.message_send_btn())
        self.mobile_function.click(ele_message_send_btn)
        time.sleep(0.2)
        self.send_message_time = time.time()

        while time.time() - self.send_message_time < 30:
            contact_photos = self.mobile_function.find_elements(AndroidMobilePageObject.contact_photo())
            if len(contact_photos) > 0 and self.mobile_function.get_rect(contact_photos[-1])["x"] < 500:
                self.get_reply_time = time.time()
                break
        if self.get_reply_time:
            cost_time = str(self.get_reply_time - self.send_message_time)
            logger.info("回复信息反应时间: " + cost_time)
            self.get_reply_time = None
            return cost_time
        else:
            logger.error("在30s内获取回复信息反应时间失败")
            return None

    def deal_with_test_data(self, data):
        """

        :param data:
        :return:
        """
        logger.info("开始处理case: " + data.case_no)

        # 发送的信息并获取回复时间
        result_reply = self.send_message_then_calculating_time_taken_to_reply(data.send_message)
        if result_reply:
            # J 给reply_cost_time_from_script赋值
            data.reply_cost_time_from_script = result_reply
        else:
            data.reply_cost_time = "在30s内获取回复信息反应时间失败"

        # 隐藏键盘
        ele_menu_btn = self.mobile_function.is_element_visible(AndroidMobilePageObject.menu_btn())
        if ele_menu_btn:
            self.mobile_function.click(ele_menu_btn)
            time.sleep(2)

        # 不管获取消息成功与否，都进行截图
        current_screenshot = PATH(os.path.join("screenshot", "case" + str(data.case_no) + ".png"))
        self.mobile_function.save_screenshot_as_png(current_screenshot)
        # H 给screenshot_from_script赋值
        data.screenshot_from_script = os.path.join("screenshot", "case" + str(data.case_no) + ".png")
        folder_name = data.link_template_screenshot_folder
        # 获取回复的信息
        if result_reply:
            latest_message = self.mobile_function.find_elements(AndroidMobilePageObject.latest_message())
            # 最多取五个消息

            flag_count = len(latest_message) if len(latest_message) < 5 else 5
            message_list = []

            folder_path = PATH(os.path.join("template", folder_name, "copy_item"))

            for i in range(len(latest_message) - 1, -1, -1):
                if flag_count > 0:
                    flag_count = flag_count - 1
                    rect = self.mobile_function.get_rect(latest_message[i])
                    x = rect["x"] + rect["width"] / 3
                    y = rect["y"] + rect["height"] / 3
                    message = self.mobile_function.get_details_message_by_copy_past_item(x=x, y=y,
                                                                                         copy_item_path=folder_path,
                                                                                         time_out=60)

                    if message:
                        if message == data.send_message:
                            break
                        message_list.append(message)
                    time.sleep(1.5)
                    latest_message = self.mobile_function.find_elements(AndroidMobilePageObject.latest_message())
                else:
                    break

            message_all = ""
            if len(message_list) > 0:
                message_list.reverse()
                message_all = "\n".join(message_list)

            # F 给reply_from_script赋值
            pattern = re.compile(r'<a.*?href="(.*?)".*?>(.*?)</a>')
            find_content_list = re.findall(pattern, message_all)

            href_link_list = []
            for find_content in find_content_list:
                # print(find_content[0], find_content[1])
                href_link_list.append(unquote(find_content[0]))
                message_all = message_all.replace(find_content[0], "")

            if len(href_link_list) > 0:
                message_all = message_all.replace('<a href="">', "").replace("</a>", "")
                # link_from_script
                data.link_from_script = "\n".join(href_link_list)

            data.reply_from_script = message_all

            # 处理回复消息中的link
            link_template_screenshot_list = data.link_template_screenshot.split("\n")
            link_screenshot_flag = 1
            link_screenshot_list = []
            if len(link_template_screenshot_list) > 0 and len(href_link_list) > 0:
                for link_template_screenshot in link_template_screenshot_list:
                    # 获取 link样例图片在原图中的坐标
                    loc = Utils.match_image_by_match_template_func(current_screenshot,
                                                                   PATH(
                                                                       os.path.join("template", folder_name,
                                                                                    link_template_screenshot)))

                    current_link_screenshot = os.path.join("screenshot", "case{}_link{}.png".format(data.case_no,
                                                                                                    link_screenshot_flag))
                    # 点击匹配到的link
                    if loc:
                        try:
                            self.mobile_function.tap(loc["x"], loc["y"])
                            # 等待页面加载
                            time.sleep(7)
                            self.mobile_function.save_screenshot_as_png(PATH(current_link_screenshot))

                            # 从link页面返回到消息界面
                            back_btn = self.mobile_function.wait_for_element_visible(AndroidMobilePageObject.back_btn())
                            self.mobile_function.click(back_btn)
                            self.mobile_function.wait_for_element_visible(AndroidMobilePageObject.title_in_chat())
                        except Exception as e:
                            logger.error(e)
                            logger.error("No link template matched for: " + current_link_screenshot)
                            current_link_screenshot = "No link template matched for: " + current_link_screenshot
                    else:
                        logger.error("No link template matched for: " + current_link_screenshot)
                        current_link_screenshot = "No link template matched for: " + current_link_screenshot

                    link_screenshot_flag = link_screenshot_flag + 1
                    link_screenshot_list.append(current_link_screenshot)

                if len(link_screenshot_list) > 0:
                    # I 给link_screenshot_from_script赋值
                    data.link_screenshot_from_script = "\n".join(link_screenshot_list)
            elif len(link_template_screenshot_list) > 0 and len(href_link_list) == 0:
                logger.error("回复消息中没有link")
                # I 给link_screenshot_from_script赋值
                data.link_screenshot_from_script = "回复消息中没有link"

    def _init_test_data(self, data):
        pass


class IOSProcess:

    def __init__(self, webdriver):
        self.driver = webdriver
        self.send_message_time = None
        self.get_reply_time = None
        self.mobile_function = MobileFunction(self.driver)

    def go_into_volkswagen_official_account(self, count_name):
        ele_address_book_btn = self.mobile_function.wait_for_element_visible(
            IOSMobilePageObject.address_book_in_home_page()
        )
        self.mobile_function.click(ele_address_book_btn)

        ele_gongzong_number_item = self.mobile_function.wait_for_element_presence(
            IOSMobilePageObject.gongzong_number_item()
        )
        rect = self.mobile_function.get_rect(ele_gongzong_number_item)
        self.mobile_function.tap(rect["x"] + rect["width"] / 2, rect["y"] + rect["height"] / 2)

        ele_search_in_gongzong_page = self.mobile_function.wait_for_element_visible(
            IOSMobilePageObject.search_in_gongzong_page()
        )
        self.mobile_function.click(ele_search_in_gongzong_page)

        ele_search_input_in_gongzong_page = self.mobile_function.wait_for_element_visible(
            IOSMobilePageObject.search_input_in_gongzong_page()
        )
        # "大众汽车金融中国测试号"
        self.mobile_function.send_text(ele_search_input_in_gongzong_page, count_name.decode("utf-8"))

        ele_target_item = self.mobile_function.wait_for_element_presence(
            IOSMobilePageObject.target_item()
        )
        rect = self.mobile_function.get_rect(ele_target_item)
        self.mobile_function.tap(rect["x"] + rect["width"] / 2, rect["y"] + rect["height"] / 2)

        # ele_title_in_chat = self.mobile_function.wait_for_element_visible(IOSMobilePageObject.title_in_chat())

    def send_message_then_calculating_time_taken_to_reply(self, value, wechat_name):
        value = value.decode("utf-8")
        logger.info("发送信息: " + value)
        # self.mobile_function.wait_for_element_visible(AndroidMobilePageObject.title_in_chat())
        time.sleep(1)
        ele_message = self.mobile_function.is_element_visible(IOSMobilePageObject.message_btn())
        if ele_message:
            self.mobile_function.click(ele_message)

        ele_message_input = self.mobile_function.wait_for_element_visible(IOSMobilePageObject.message_input())
        self.mobile_function.send_text(ele_message_input, value)
        ele_message_send_btn = self.mobile_function.wait_for_element_visible(IOSMobilePageObject.message_send_btn())
        self.mobile_function.click(ele_message_send_btn)
        time.sleep(0.2)
        self.send_message_time = time.time()

        while time.time() - self.send_message_time < 30:
            latest_message = self.mobile_function.wait_for_element_presence(IOSMobilePageObject.latest_message())
            text = self.mobile_function.get_text(latest_message)
            if text.startswith(wechat_name) or text.startswith("该公众号提供的服务出现故障".decode("utf-8")):
                self.get_reply_time = time.time()
                break
        if self.get_reply_time:
            cost_time = str(self.get_reply_time - self.send_message_time)
            logger.info("取回复信息反应时间: " + cost_time)
            self.get_reply_time = None
            return cost_time
        else:
            logger.error("在30s内获取回复信息反应时间失败")
            return None

    def deal_with_test_data(self, data, wechat_name):
        """

        :param data:
        :return:
        """
        logger.info("开始处理case: " + data.case_no)

        # 发送的信息并获取回复时间
        result_reply = self.send_message_then_calculating_time_taken_to_reply(data.send_message, wechat_name)
        if result_reply:
            # J 给reply_cost_time_from_script赋值
            data.reply_cost_time_from_script = result_reply
        else:
            data.reply_cost_time = "在30s内获取回复信息反应时间失败"

        # 隐藏键盘
        ele_menu_btn = self.mobile_function.is_element_visible(IOSMobilePageObject.menu_btn())
        if ele_menu_btn:
            self.mobile_function.click(ele_menu_btn)
            time.sleep(2)

        # 不管获取消息成功与否，都进行截图
        current_screenshot = PATH(os.path.join("screenshot", "case" + str(data.case_no) + ".png"))
        self.mobile_function.save_screenshot_as_png(current_screenshot)
        # H 给screenshot_from_script赋值
        data.screenshot_from_script = os.path.join("screenshot", "case" + str(data.case_no) + ".png")
        folder_name = data.link_template_screenshot_folder

        # 获取回复的信息
        if result_reply:
            self.mobile_function.wait_for_element_presence(IOSMobilePageObject.latest_message())
            all_message_eles = self.mobile_function.find_elements(IOSMobilePageObject.all_message())

            flag_count = len(all_message_eles) if len(all_message_eles) < 8 else 8
            message_list = []

            for i in range(len(all_message_eles) - 1, -1, -1):
                if flag_count > 0:
                    flag_count = flag_count - 1
                    message = self.mobile_function.get_text(all_message_eles[i])
                    logger.info("Get the message: " + message)
                    if message.endswith(data.send_message):
                        break
                    message = message.replace(wechat_name + "说".decode("utf-8"), "")
                    if message not in message_list:
                        message_list.append(message)

            message_list.reverse()
            message = "\n".join(message_list)
            # F 给reply_from_script赋值
            pattern = re.compile(r'<a.*?href="(.*?)".*?>(.*?)</a>')
            find_content_list = re.findall(pattern, message)

            href_link_list = []
            for find_content in find_content_list:
                # print(find_content[0], find_content[1])
                href_link_list.append(unquote(find_content[0]))
                message = message.replace(find_content[0], "")

            if len(href_link_list) > 0:
                message = message.replace('<a href="">', "").replace("</a>", "")
                # link_from_script
                data.link_from_script = "\n".join(href_link_list)

            data.reply_from_script = message

            # 处理回复消息中的link
            link_template_screenshot_list = data.link_template_screenshot.split("\n")
            link_screenshot_flag = 1
            link_screenshot_list = []
            if len(link_template_screenshot_list) > 0 and len(href_link_list) > 0:
                for link_template_screenshot in link_template_screenshot_list:
                    # 获取 link样例图片在原图中的坐标
                    loc = Utils.match_image_by_match_template_func(current_screenshot,
                                                                   PATH(os.path.join("template", folder_name,
                                                                                     link_template_screenshot))
                                                                   , self.driver.get_window_rect())

                    current_link_screenshot = os.path.join("screenshot", "case{}_link{}.png".format(data.case_no,
                                                                                                    link_screenshot_flag))
                    # 点击匹配到的link
                    if loc:
                        try:
                            self.mobile_function.tap(loc["x"], loc["y"])
                            # todo: 等待页面加载
                            time.sleep(7)
                            self.mobile_function.save_screenshot_as_png(PATH(current_link_screenshot))

                            # 从link页面返回到消息界面
                            back_btn = self.mobile_function.wait_for_element_visible(IOSMobilePageObject.back_btn())
                            self.mobile_function.click(back_btn)
                            self.mobile_function.wait_for_element_visible(IOSMobilePageObject.title_in_chat())
                        except Exception as e:
                            logger.error(e)
                            logger.error("No link template matched for: " + current_link_screenshot)
                            current_link_screenshot = "No link template matched for: " + current_link_screenshot
                    else:
                        logger.error("No link template matched for: " + current_link_screenshot)
                        current_link_screenshot = "No link template matched for: " + current_link_screenshot

                    link_screenshot_flag = link_screenshot_flag + 1
                    link_screenshot_list.append(current_link_screenshot)

                if len(link_screenshot_list) > 0:
                    # I 给link_screenshot_from_script赋值
                    data.link_screenshot_from_script = "\n".join(link_screenshot_list)
            elif len(link_template_screenshot_list) > 0 and len(href_link_list) == 0:
                logger.error("回复消息中没有link")
                # I 给link_screenshot_from_script赋值
                data.link_screenshot_from_script = "回复消息中没有link"


class CaseDataModel:

    def __init__(self, case_no, send_message, reply, link_template_screenshot_folder, link_template_screenshot,
                 reply_from_script, link_from_script, screenshot_from_script, link_screenshot_from_script,
                 reply_cost_time_from_script, result):
        self.case_no = str(case_no)
        self.send_message = send_message
        self.reply = reply
        self.link_template_screenshot_folder = link_template_screenshot_folder
        self.link_template_screenshot = link_template_screenshot
        self.reply_from_script = reply_from_script
        self.link_from_script = link_from_script
        self.screenshot_from_script = screenshot_from_script
        self.link_screenshot_from_script = link_screenshot_from_script
        self.reply_cost_time_from_script = reply_cost_time_from_script
        self.result = result


def clean_data():
    del_list = os.listdir(PATH("screenshot"))
    for f in del_list:
        file_path = os.path.join(PATH("screenshot"), f)
        os.remove(file_path)


def android_steps(test_data_list, wechat_name):
    desired_caps_android_wechat = {
        "platformName": "Android",
        "platformVersion": os.getenv("APPIUM_DEVICE_VERSION", "10"),
        "automationName": os.getenv("APPIUM_AUTOMATION_NAME", "Appium"),
        "appActivity": os.getenv("APPIUM_APP_ACTIVITY", "com.tencent.mm.ui.LauncherUI"),
        "appPackage": os.getenv("APPIUM_APP_PACKAGE", "com.tencent.mm"),
        "deviceName": os.getenv("APPIUM_DEVICE_NAME", "AKC7N18907000186"),
        "newCommandTimeout": 7200,
        "noReset": True,
        'chromeOptions': {'androidProcess': 'com.tencent.mm:tools'}
    }

    # 1. 从excel读取数据
    test_data_list_copy = test_data_list[1:]
    logger.info("Android case 统计: " + str(len(test_data_list_copy)))

    driver = webdriver.Remote(os.getenv("APPIUM_URL", 'http://localhost:4723/wd/hub'), desired_caps_android_wechat)
    android_process = AndroidProcess(driver)

    # 2. 进入公众号
    android_process.go_into_volkswagen_official_account(wechat_name)

    # 3. 处理消息
    for test_data in test_data_list_copy:
        try:
            # 需要判断是否回到主界面
            try:
                android_process.mobile_function.wait_for_element_visible(AndroidMobilePageObject.title_in_chat())
            except Exception as e:
                try:
                    driver.close_app()
                    time.sleep(2)
                    driver.launch_app()
                    android_process.go_into_volkswagen_official_account(wechat_name)
                except Exception as e:
                    logger.warn(e)
                    retry_count = 5
                    while retry_count > 0:
                        logger.info("需要等待10s,然后重启driver")
                        time.sleep(10)
                        try:
                            try:
                                driver.quit()
                            except Exception:
                                pass
                            time.sleep(2)
                            driver = webdriver.Remote(os.getenv("APPIUM_URL", 'http://localhost:4723/wd/hub'), desired_caps_android_wechat)
                            android_process = AndroidProcess(driver)
                            android_process.go_into_volkswagen_official_account(wechat_name)
                            break
                        except Exception as e:
                            logger.warn(e)
                        retry_count = retry_count - 1

            android_process.deal_with_test_data(test_data)
        except Exception as e:
            logger.warn(e)

    # 数据写进excel
    Utils.write_data_into_excel(PATH("test_case_example.xlsx"), test_data_list_copy)
    logger.info("write_data_into_excel")


def ios_steps(test_data_list, wechat_name):
    desired_caps_ios_wechat = {
        "platformName": "iOS",
        "PlatformVersion": os.getenv('APP_DEVICE_VERSION', "12.2"),
        "deviceName": os.getenv('APP_DEVICE_NAME', "iPhone"),
        "automationName": "XCUITest",
        "udid": os.getenv("APP_UDID", "029d553ea04ba899509dc0630fda19bdac61231a"),
        "bundleId": os.getenv("APP_BUNDLEIDENTIFIER", "com.tencent.xin"),
        "newCommandTimeout": 7200,
        "startIWDP": True,
        "webDriverAgentUrl": os.getenv("WEBDRIVERAGENT_URL", "http://localhost:8100")
    }

    # 1. 从excel读取数据
    test_data_list_copy = test_data_list[1:]
    logger.info("IOS case 统计: " + str(len(test_data_list_copy)))

    driver = webdriver.Remote(os.getenv("APPIUM_URL", 'http://localhost:4723/wd/hub'), desired_caps_ios_wechat)
    ios_process = IOSProcess(driver)

    # 2. 进入公众号
    ios_process.go_into_volkswagen_official_account(wechat_name)

    # 3. 处理消息
    for test_data in test_data_list_copy:
        try:
            # 需要判断是否回到主界面
            try:
                ios_process.mobile_function.wait_for_element_visible(IOSMobilePageObject.title_in_chat())
            except Exception as e:
                try:
                    driver.close_app()
                    time.sleep(2)
                    driver.launch_app()
                    ios_process.go_into_volkswagen_official_account(wechat_name)
                except Exception as e:
                    logger.warn(e)
                    retry_count = 5
                    while retry_count > 0:
                        logger.info("需要等待10s,然后重启driver")
                        time.sleep(10)
                        try:
                            try:
                                driver.quit()
                            except Exception:
                                pass
                            time.sleep(2)
                            driver = webdriver.Remote(os.getenv("APPIUM_URL", 'http://localhost:4723/wd/hub'), desired_caps_ios_wechat)
                            ios_process.go_into_volkswagen_official_account(wechat_name)
                            ios_process = IOSProcess(driver)
                            break
                        except Exception as e:
                            logger.warn(e)
                        retry_count = retry_count - 1

            ios_process.deal_with_test_data(test_data, wechat_name)
        except Exception as e:
            logger.info(e)

    # 数据写进excel
    Utils.write_data_into_excel(PATH("test_case_example.xlsx"), test_data_list_copy)
    logger.info("write_data_into_excel")


if __name__ == '__main__':
    clean_data()
    # 1. 从excel读取数据
    test_data_dict = Utils.load_data_from_excel(PATH("test_case_example.xlsx"))
    config_data = test_data_dict["config_data"]
    logger.info("配置信息为: os = {}, 公众号 = {}".format(config_data["os"], config_data["name"]))

    os_name = config_data["os"].lower()

    if os_name == "android":
        android_steps(test_data_dict["test_data"], config_data["name"])
    elif os_name == "ios":
        ios_steps(test_data_dict["test_data"], config_data["name"])
    else:
        logger.error("请正确配置excel中config sheet中系统信息")

    logger.info("完成测试: ")
